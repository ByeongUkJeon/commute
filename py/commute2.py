import os
import pickle
import sys
import shutil
import time
import pandas as pd
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5 import QtCore, QtTest
from PyQt5.QtGui import *
from threading import Thread
from distutils.dir_util import copy_tree
import calendar
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
import openpyxl
import datetime as dt
import subprocess
days = ['월', '화', '수', '목', '금', '토', '일']

class Thread(QThread):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.dttime = dt.datetime.now()
        self.timer = QTimer(self)
        self.timer.setInterval(1000)
        self.timer.timeout.connect(self.run)
        self.timer.start()

    def run(self):
        localtime = time.localtime()
        rTimel = time.strftime("%M:%S", localtime)
        ttime = QTime.currentTime().toString("AP hh:mm:ss")
        self.parent.timeLabel.setText(ttime)
        if rTimel == self.dttime.strftime("10:00") or rTimel == self.dttime.strftime("40:00"):
         #    os.system(r'commute/commute.exe')
        #    QCoreApplication.instance().quit()
            self.parent.update()
        QApplication.processEvents()

class Thread2(QThread):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent

    def run(self):
        QApplication.processEvents()
        inp = self.parent.lineEdit1.text()
        inp = inp[0:6]
        self.parent.lineEdit1.setText('')

        f = open(r'data.JEON', 'rb')
        data = pickle.load(f)
        f.close()

        #self.commute(  inp, data)
        # self.privateCommute(inp, data)
        # self.publicCommute(inp, data)
        QApplication.processEvents()
        hour = dt.datetime.now()
        datetime = QDateTime.currentDateTime()
        tim = QTime.currentTime()
        today1pm = hour.replace(hour=13, minute=0, second=0, microsecond=0)
        orignPath = os.getcwd() + '\\' + datetime.toString('yyyy') + '년\\' + datetime.toString('MM') + '월'
        filePath = orignPath + '\\'
        if inp in data:
            name = data[inp][0]

            privatePath = filePath + name + '\\'
            privateFileName = '출퇴근기록부_' + name + '_' + datetime.toString('MM') + '월' + '.xlsx'

            if not os.path.isdir(privatePath):
                os.makedirs(privatePath)
                df = pd.DataFrame([int(inp), data[inp][0], data[inp][1],
                                   int(datetime.toString('yyyy') + datetime.toString('MM') + datetime.toString('dd')),
                                   datetime.toString('ddd')])
                dffirstDay = pd.DataFrame([int(inp), data[inp][0], data[inp][1]])

                try:

                    wb = openpyxl.load_workbook(r'template.xlsx')
                    ws = wb.active
                    ws.print_area = "A1:L42"
                    lastDay = calendar.monthrange(int(datetime.toString('yyyy')), int(datetime.toString('M')))[1]
                    ws.merge_cells('K2:L2')
                    ws['K2'] = datetime.toString('yyyy') + '.' + datetime.toString('MM') + '.01~' + datetime.toString(
                        'yyyy') + '.' + datetime.toString('MM') + '.' + str(lastDay)
                    ws['K2'].font = Font(size=11, bold=True)
                    ws['K2'].alignment = Alignment(horizontal='center', vertical='center')
                    ws['A1'].font = Font(size=24, bold=True)
                    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                    for i in range(5, lastDay + 5):
                        y = 2
                        for r in dataframe_to_rows(dffirstDay, index=False, header=False):
                            ws.cell(i, y).value = r[0]
                            if(i - 4 < 10):
                                date = '0' + str(i - 4)
                            else:
                                date = str(i - 4)
                            y += 1
                        day = days[dt.date(int(datetime.toString('yyyy')), int(datetime.toString('M')), i - 4).weekday()]
                        ws.cell(i, 6).value = day
                        ws.cell(i, 5).value = int(datetime.toString('yyyy') + datetime.toString('MM') + date)

                    if hour > today1pm:  # 1시 이후
                        ws.cell(int(datetime.toString('dd')) + 4, 8).value = tim.toString('hhmmss')
                        self.parent.label.setText(data[inp][0] + '님 안녕히가세요.')
                    else:  # 1시 이전
                        ws.cell(int(datetime.toString('dd')) + 4, 7).value = tim.toString('hhmmss')
                        self.parent.label.setText(data[inp][0] + '님 환영합니다.')

                    wb.save(privatePath + privateFileName)
                    # copy_tree(privatePath, copyPath + name + '\\')
                    return
                except:
                    self.parent.label.setText('오류1 발생')
                    return
            else:
                if not (os.path.isfile(privatePath + privateFileName)):
                    df = pd.DataFrame([int(inp), data[inp][0], data[inp][1],
                                       int(datetime.toString('yyyy') + datetime.toString('MM') + datetime.toString(
                                           'dd')),
                                       datetime.toString('ddd')])
                    dffirstDay = pd.DataFrame([int(inp), data[inp][0], data[inp][1]])

                    try:
                        wb = openpyxl.load_workbook(r'template.xlsx')
                        ws = wb.active
                        ws.print_area = "A1:L42"

                        lastDay = calendar.monthrange(int(datetime.toString('yyyy')), int(datetime.toString('M')))[1]
                        ws.merge_cells('K2:L2')
                        ws['K2'] = datetime.toString('yyyy') + '.' + datetime.toString(
                            'MM') + '.01~' + datetime.toString(
                            'yyyy') + '.' + datetime.toString('MM') + '.' + str(lastDay)
                        ws['K2'].font = Font(size=11, bold=True)
                        ws['K2'].alignment = Alignment(horizontal='center', vertical='center')
                        ws['A1'].font = Font(size=24, bold=True)
                        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

                        for i in range(5, lastDay + 5):
                            y = 2
                            for r in dataframe_to_rows(dffirstDay, index=False, header=False):
                                ws.cell(i, y).value = r[0]
                                if (i - 4 < 10):
                                    date = '0' + str(i - 4)
                                else:
                                    date = str(i - 4)
                                y += 1
                            day = days[
                                dt.date(int(datetime.toString('yyyy')), int(datetime.toString('M')), i - 4).weekday()]
                            ws.cell(i, 6).value = day
                            ws.cell(i, 5).value = int(datetime.toString('yyyy') + datetime.toString('MM') + date)
                        if hour > today1pm:  # 1시 이후
                            ws.cell(int(datetime.toString('dd')) + 4, 8).value = tim.toString('hhmmss')
                            self.parent.label.setText(data[inp][0] + '님 안녕히가세요.')
                        else:  # 1시 이전
                            ws.cell(int(datetime.toString('dd')) + 4, 7).value = tim.toString('hhmmss')
                            self.parent.label.setText(data[inp][0] + '님 환영합니다.')
                        wb.save(privatePath + privateFileName)
                        # copy_tree(privatePath, copyPath + name + '\\')
                        return
                    except:
                        self.parent.label.setText("오류2 발생")
                        return

                if (hour > today1pm):  # 개인, 퇴근
                    wb = openpyxl.load_workbook(privatePath + privateFileName)
                    ws = wb.active
                    ws.print_area = "A1:L42"

                    lastDay = calendar.monthrange(int(datetime.toString('yyyy')), int(datetime.toString('M')))[1]
                    try:

                        ws.cell(int(datetime.toString('dd')) + 4, 8).value = tim.toString('hhmmss')
                        wb.save(privatePath + privateFileName)
                        # copy_tree(privatePath, copyPath + name + '\\')
                        self.parent.label.setText(data[inp][0] + '님 안녕히가세요 :)')
                        return
                    except PermissionError:
                        self.parent.label.setText('접근 오류 발생')
                        return
                    except:
                        self.parent.label.setText('오류3 발생')
                        return
                else:  # 개인 출근
                    wb = openpyxl.load_workbook(privatePath + privateFileName)
                    ws = wb.active
                    try:
                        if ws.cell(int(datetime.toString('dd')) + 4, 7).value is not None:
                            self.parent.label.setText(data[inp][0] + '님 이미 출근 처리 되었습니다.')
                            return

                        ws.cell(int(datetime.toString('dd')) + 4, 7).value = tim.toString('hhmmss')
                        wb.save(privatePath + privateFileName)
                        #copy_tree(privatePath, copyPath + name + '\\')
                        self.parent.label.setText(data[inp][0] + '님 환영합니다. :)')
                        return
                    except PermissionError:
                        self.parent.label.setText('접근 오류 발생')
                        return
                    except:
                        self.parent.label.setText('오류4 발생')
                        return
        else:
            self.parent.label.setText("존재하지 않는 사용자입니다.")
            return



class MyWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.timeLabel = QLabel()
        self.timeLabel.setFont(QFont("굴림", 15))
        self.x = Thread(self)
        self.x.start()
        self.y = Thread2(self)
        self.initUI()

    def initUI(self):
        os.chdir('../')
        self.tableDialog = QDialog()
        self.tableDialog.resize(450, 600)
        self.tableDialog.setWindowTitle("사용자 목록")
        self.tableDialog.tableWidget = QTableWidget()
        self.tableDialog.tableWidget.setColumnCount(3)
        self.tableDialog.deleteButton = QPushButton("삭제")
        self.tableDialog.tableWidget.setHorizontalHeaderLabels(['바코드', '이름', '직위'])
        self.tableDialog.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tableDialog.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableDialog.layout = QVBoxLayout()
        self.tableDialog.layout.addWidget(self.tableDialog.tableWidget)
        self.tableDialog.layout.addWidget(self.tableDialog.deleteButton)
        self.tableDialog.setLayout(self.tableDialog.layout)
        self.tableDialog.tableWidget.cellPressed.connect(self.rowReturn)
        self.tableDialog.deleteButton.clicked.connect(self.deleteCell)

        self.row = -1
        self.resize(600, 350)
        self.center()
        self.label = QLabel("안녕하세요.")
        self.lineEdit1 = QLineEdit()
        self.managerButton = QPushButton('관리')


        self.managerButton.clicked.connect(self.openManagerDialog)
        self.label.setFont(QFont("굴림", 16))

        self.setWindowTitle("출퇴근 프로그램 20211021")


        self.lineEdit1.returnPressed.connect(self.y.start)

        self.dialog = QDialog()
        self.dialog.resize(500, 300)
        self.dialog.setWindowTitle("사용자 등록")
        self.dialog.nameLabel = QLabel("이름")
        self.dialog.idLabel = QLabel("바코드 번호")
        self.dialog.rankLabel = QLabel("직위")
        self.dialog.id = QLineEdit()
        self.dialog.name = QLineEdit()
        self.dialog.rank = QLineEdit()

        self.managerDialog = QDialog()
        self.managerDialog.resize(300, 500)
        self.managerDialog.setWindowTitle("관리 메뉴")
        self.managerDialog.addButton = QPushButton("등록")
        self.managerDialog.deleteButton = QPushButton("삭제")
        self.managerDialog.syncButton = QPushButton("동기화")

        self.syncDialog = QDialog()
        self.syncDialog.resize(500, 300)
        self.syncDialog.id = QLineEdit()
        self.syncDialog.nameLabel = QLabel("이름")
        self.syncDialog.idLabel = QLabel("바코드 번호")
        self.syncDialog.name = QLineEdit()
        self.syncDialog.setWindowTitle("동기화")

        layout = QVBoxLayout()
        secondLayout = QHBoxLayout()
        layout.addStretch(2)
        layout.addWidget(self.timeLabel)
        layout.addStretch(2)
        layout.addWidget(self.label)
        layout.addStretch(2)
        layout.addWidget(self.lineEdit1)
        layout.addStretch(1)
        secondLayout.addWidget(self.managerButton)
        self.setLayout(layout)
        layout.addLayout(secondLayout)
        self.show()

    def openManagerDialog(self):
        self.setDisabled(True)
        mLayout = QGridLayout()
        mLayout.addWidget(self.managerDialog.addButton)
        mLayout.addWidget(self.managerDialog.deleteButton)
        #mLayout.addWidget(self.managerDialog.syncButton)
        self.managerDialog.addButton.clicked.connect(self.openAddUserDialog)
        self.managerDialog.deleteButton.clicked.connect(self.test)
        #self.managerDialog.syncButton.clicked.connect(self.openSyncDialog)

        self.managerDialog.open()
        self.managerDialog.setLayout(mLayout)
        self.managerDialog.closeEvent = self.closeEvent


    def test(self):
        fi = open(r"data.JEON", 'rb')
        deleteData = pickle.load(fi)
        fi.close()
        self.tableDialog.tableWidget.setRowCount(len(deleteData))

        # self.tableDialog.tableWidget.insertRow(self.tableDialog.tableWidget.rowCount())
        # self.tableDialog.tableWidget.setItem(self.tableDialog.tableWidget.rowCount()-1, 0, QTableWidgetItem("51251232"))
        for key, value in enumerate(deleteData.items()):
            self.tableDialog.tableWidget.setItem(key, 0, QTableWidgetItem(str(value[0])))
            self.tableDialog.tableWidget.setItem(key, 1, QTableWidgetItem(str(value[1][0])))
            self.tableDialog.tableWidget.setItem(key, 2, QTableWidgetItem(str(value[1][1])))

        self.tableDialog.open()


    def rowReturn(self, row):
        self.row = row

    def deleteCell(self):
        if (self.row == -1):
            return
        barcode = self.tableDialog.tableWidget.item(self.row, 0).text()
        f = open(r"data.JEON", 'rb')
        deta = pickle.load(f)
        f.close()
        del deta[barcode]
        f = open(r"data.JEON", 'wb')
        pickle.dump(deta, f)
        f.close()
        self.tableDialog.tableWidget.removeRow(self.row)
        self.row = -1
        return
    def openSyncDialog(self):
        self.syncDialog.name.setText('')
        self.syncDialog.sync = QPushButton("동기화")
        self.syncDialog.cancel = QPushButton("취소")
        self.syncDialog.layout = QVBoxLayout()
        self.syncDialog.cancel.setAutoDefault(False)
        self.syncDialog.sync.setAutoDefault(False)
        self.syncDialog.layout.addStretch(3)
        self.syncDialog.layout.addWidget(self.syncDialog.idLabel)
        self.syncDialog.layout.addStretch(1)
        self.syncDialog.layout.addWidget(self.syncDialog.id)
        self.syncDialog.layout.addStretch(3)
        self.syncDialog.secondLayout = QHBoxLayout()
        self.syncDialog.secondLayout.addStretch(1)
        self.syncDialog.secondLayout.addWidget(self.syncDialog.sync)
        self.syncDialog.secondLayout.addStretch(1)
        self.syncDialog.secondLayout.addWidget(self.syncDialog.cancel)
        self.syncDialog.secondLayout.addStretch(1)
        self.syncDialog.layout.addLayout(self.syncDialog.secondLayout)
        self.syncDialog.id.returnPressed.connect(lambda: self.namesetFocus(self.syncDialog))
        self.syncDialog.sync.clicked.connect(lambda: self.syncUser(self.syncDialog.id.text()[0:6]))

        self.syncDialog.cancel.clicked.connect(lambda: self.dialog_close(self.syncDialog))
        self.syncDialog.setLayout(self.syncDialog.layout)
        self.syncDialog.id.setFocus()
        self.syncDialog.id.setCursorPosition(0)
        self.syncDialog.open()
    def syncUser(self, id):
        if not id:
            self.errorMessageOpen("바코드 입력 에러")
            return
        if not id.isdigit():
            self.errorMessageOpen("바코드는 숫자만 가능합니다.")
            return

    def openAddUserDialog(self):
        self.setDisabled(True)
        self.dialog.id.setText('')

        self.dialog.name.setText('')
        self.dialog.rank.setText('')
        self.dialog.add = QPushButton("등록")
        self.dialog.cancel = QPushButton("취소")
        self.dialog.layout = QVBoxLayout()
        self.dialog.cancel.setAutoDefault(False)
        self.dialog.add.setAutoDefault(False)
        self.dialog.layout.addStretch(3)
        self.dialog.layout.addWidget(self.dialog.idLabel)
        self.dialog.layout.addStretch(3)
        self.dialog.layout.addWidget(self.dialog.id)
        self.dialog.layout.addStretch(9)
        self.dialog.layout.addWidget(self.dialog.nameLabel)
        self.dialog.layout.addStretch(3)
        self.dialog.layout.addWidget(self.dialog.name)
        self.dialog.layout.addStretch(9)
        self.dialog.layout.addWidget(self.dialog.rankLabel)
        self.dialog.layout.addStretch(3)
        self.dialog.layout.addWidget(self.dialog.rank)
        self.dialog.layout.addStretch(9)
        self.dialog.secondLayout = QHBoxLayout()
        self.dialog.secondLayout.addStretch(1)
        self.dialog.secondLayout.addWidget(self.dialog.add)
        self.dialog.secondLayout.addStretch(1)
        self.dialog.secondLayout.addWidget(self.dialog.cancel)
        self.dialog.secondLayout.addStretch(1)
        self.dialog.layout.addLayout(self.dialog.secondLayout)
        self.dialog.id.returnPressed.connect(lambda: self.namesetFocus(self.dialog))
        self.dialog.add.clicked.connect(lambda: self.addUser(self.dialog.id.text()[0:6], self.dialog.name.text(), self.dialog.rank.text()))

        self.dialog.cancel.clicked.connect(lambda: self.dialog_close(self.dialog))
        self.dialog.setLayout(self.dialog.layout)
        self.dialog.id.setFocus()
        self.dialog.id.setCursorPosition(0)
        self.dialog.open()

    def namesetFocus(self, dial):
        dial.name.setFocus()
        dial.id.setText(dial.id.text()[0:6])
    def errorMessageOpen(self, error):
        QMessageBox.warning(self, '에러 발생', error)
        self.dialog.close()

    def showMessageBox(self, id, name, rank):
        QMessageBox.information(self, '등록 완료', '바코드 : ' + id + '\n이름 : ' + name + '\n직위 : ' + rank + '\n등록완료')
        self.dialog.close()
        self.managerDialog.close()
    def closeEvent(self, event):
        self.setDisabled(False)

    def dialog_close(self, dial):
        dial.close()

    def addUser(self, id, name, rank):
        if not id:
            self.errorMessageOpen("바코드 입력 에러")
            self.managerDialog.close()
            self.setDisabled(False)
            return
        if not id.isdigit():
            self.errorMessageOpen("바코드는 숫자만 가능합니다.")
            self.managerDialog.close()
            self.setDisabled(False)
            return
        if not name:
            self.errorMessageOpen("이름 입력 에러")
            self.managerDialog.close()
            self.setDisabled(False)
            return
        if name.isdigit():
            self.errorMessageOpen("이름은 문자열만 가능합니다")
            self.managerDialog.close()
            self.setDisabled(False)
            return
        if not rank:
            self.errorMessageOpen("직위 입력 에러")
            self.managerDialog.close()
            self.setDisabled(False)
            return
        if rank.isdigit():
            self.errorMessageOpen("직위는 문자열만 가능합니다")
            self.managerDialog.close()
            self.setDisabled(False)
            return

        addf = open(r"data.JEON", 'rb')
        addData = pickle.load(addf)
        addf.close()
        addData[id] = [name, rank]
        addf = open(r"data.JEON", 'wb')
        pickle.dump(addData, addf)
        self.showMessageBox(id, name, rank)
        addf.close()
        return

    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())


if __name__ == "__main__":
    app = QApplication(sys.argv)
    mywindow = MyWindow()
    mywindow.show()
    app.exec_()